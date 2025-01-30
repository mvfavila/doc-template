import os
import sys
import re
from docx import Document
from docx2pdf import convert
from pathlib import Path
from openpyxl import load_workbook

# Get the directory of the executable
EXECUTABLE_DIR = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent

# Function to find a single file with a specific extension
def find_single_file(extension):
    files = [f for f in EXECUTABLE_DIR.iterdir() if f.suffix == extension]
    if len(files) == 0:
        print(f"Erro: Nenhum arquivo '{extension}' encontrado no diretório {EXECUTABLE_DIR}.")
        input("Pressione ENTER para fechar")
        sys.exit(1)
    if len(files) > 1:
        print(f"Erro: Mais de um arquivo '{extension}' encontrado no diretório {EXECUTABLE_DIR}. Certifique-se de que haja apenas um.")
        input("Pressione ENTER para fechar")
        sys.exit(1)
    return files[0]

# Function to validate the .xlsx file and check for empty cells
def validate_xlsx(xlsx_file):
    xlsx_path = EXECUTABLE_DIR / xlsx_file
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active

    # Ensure the 'NUMERO_DO_PROCESSO' column exists
    header = [cell.value for cell in sheet[1]]
    if "NUMERO_DO_PROCESSO" not in header:
        print("Erro: Uma das colunas do arquivo XLSX deve ser nomeada como 'NUMERO_DO_PROCESSO'.")
        input("Pressione ENTER para fechar")
        sys.exit(1)

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start at row 2 to skip the header
        row_data = {header[i]: row[i] for i in range(len(row))}
        rows.append(row_data)

        # Check for empty cells in the row
        for column_name, value in row_data.items():
            if value is None or str(value).strip() == "":
                print(f"Erro: A célula na linha {sheet.iter_rows(min_row=2, values_only=True).index(row) + 2}, coluna '{column_name}' está vazia.")
                input("Pressione ENTER para fechar")
                sys.exit(1)
    
    return rows

# Function to sanitize file names
def sanitize_filename(filename):
    return re.sub(r'[^\w\-_\.]', '_', filename)

# Function to generate PDF from DOCX
def generate_pdf(docx_file, pdf_file):
    try:
        docx_path = EXECUTABLE_DIR / docx_file
        convert(str(docx_path))  # In-place conversion
        pdf_path = docx_path.with_suffix(".pdf").resolve()
        if pdf_path.exists():
            print(f"PDF gerado com sucesso: {pdf_path}")
        else:
            print(f"PDF não foi gerado: {docx_file}")
    except Exception as e:
        print(f"Erro durante geração do PDF {docx_file}: {e}")
        input("Pressione ENTER para fechar")

# Main logic
def main():
    print("Como usar")
    print("1. Ao iniciar, deve existir exatamente um arquivo .docx e um arquivo .xlsx no mesmo diretório desse executável.")
    print("2. O arquivo .docx deve possuir um placeholder para cada coluna do arquivo .xlsx. ex.: {{NUMERO_DO_PROCESSO}}")
    print("3. Certifique-se de que existe uma coluna chamada 'NUMERO_DO_PROCESSO' no arquivo .xlsx.")
    print("4. Os nomes das colunas do arquivo .xlsx devem ser exatamente as mesmas dos placeholders do arquivo .docx.")
    print()

    print("Processando...")
    # Find the .docx template file
    template_file_path = find_single_file('.docx')
    print(f"Arquivo de template encontrado: {template_file_path.name}")

    # Find the .xlsx data file
    xlsx_file_path = find_single_file('.xlsx')
    print(f"Arquivo XLSX encontrado: {xlsx_file_path.name}")

    # Validate and read the XLSX file
    rows = validate_xlsx(xlsx_file_path)

    # Process each row in the XLSX file
    for row in rows:
        process_number = sanitize_filename(row["NUMERO_DO_PROCESSO"])
        docx_output = EXECUTABLE_DIR / f"{process_number}_{template_file_path.name}"
        pdf_output = EXECUTABLE_DIR / f"{process_number}_{template_file_path.stem}.pdf"

        # Create a copy of the template
        doc = Document(template_file_path)

        # Replace placeholders with values from the row
        for paragraph in doc.paragraphs:
            for placeholder, value in row.items():
                if f"{{{{{placeholder}}}}}" in paragraph.text:
                    paragraph.text = paragraph.text.replace(f"{{{{{placeholder}}}}}", str(value))

        # Save the filled .docx file
        doc.save(docx_output)

        # Generate the corresponding PDF file
        generate_pdf(docx_output, pdf_output)
        print(f"Arquivos gerados: {docx_output} e {pdf_output}")

    input("Pressione ENTER para fechar")

if __name__ == "__main__":
    main()
